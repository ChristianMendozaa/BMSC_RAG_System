import logging
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class TextBlock:
    text: str
    page_number: int | None
    block_type: str = "text"


@dataclass
class ParseResult:
    text_blocks: list[TextBlock] = field(default_factory=list)
    image_blocks: list = field(default_factory=list)


def parse(file_bytes: bytes) -> ParseResult:
    result = ParseResult()

    try:
        # read_only no expone imágenes/gráficos embebidos: solo se extrae texto de celdas
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Failed to open XLSX: %s", exc)
        raise

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        try:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows_text.append("\t".join(cells))

            if rows_text:
                content = f"[Hoja: {sheet_name}]\n" + "\n".join(rows_text)
                result.text_blocks.append(TextBlock(text=content, page_number=sheet_idx))
        except Exception as exc:
            logger.warning("Skipping XLSX sheet '%s': %s", sheet_name, exc)

    wb.close()
    return result
